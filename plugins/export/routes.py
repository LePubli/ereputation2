"""Plugin Export — Génération Excel/CSV des prospects."""
import csv
import io
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from loguru import logger
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from core.auth import CurrentUser
from core.database import get_db

router = APIRouter()

ACCENT = {"green": "3FB950", "blue": "2F81F7", "orange": "D29922", "red": "F85149"}


def _score_color(score):
    if score is None: return "8B949E"
    if score >= 75: return ACCENT["green"]
    if score >= 50: return ACCENT["blue"]
    if score >= 25: return ACCENT["orange"]
    return ACCENT["red"]


@router.get("/prospects.xlsx")
async def export_prospects_excel(
    ids: Optional[List[str]] = Query(None),
    region: Optional[str] = Query(None),
    score_min: Optional[int] = Query(None),
    limit: int = Query(5000, le=10000),
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        where, params = [], {"limit": limit}
        if ids:
            ph = ", ".join(f":id{i}" for i, _ in enumerate(ids))
            where.append(f"p.id IN ({ph})")
            for i, v in enumerate(ids): params[f"id{i}"] = v
        if region:
            where.append("p.region = :region"); params["region"] = region
        if score_min:
            where.append("p.propensity_score >= :sm"); params["sm"] = score_min

        where_sql = "WHERE " + " AND ".join(where) if where else ""
        rows = (await db.execute(text(f"""
            SELECT p.company_name, p.siren, p.city, p.region, p.naf_code, p.naf_label,
                   p.employee_range, p.phone, p.email, p.website, p.propensity_score,
                   COALESCE(ps.name, 'Nouveau'), p.created_at
            FROM prospects p
            LEFT JOIN pipeline_stages ps ON p.stage_id = ps.id
            {where_sql}
            ORDER BY p.propensity_score DESC NULLS LAST, p.company_name
            LIMIT :limit
        """), params)).fetchall()

        wb = Workbook(); ws = wb.active; ws.title = "Prospects"
        ws.sheet_view.showGridLines = False

        headers = ["Entreprise","SIREN","Ville","Région","Code NAF","Secteur",
                   "Effectif","Téléphone","Email","Site web","Score","Étape","Créé le"]
        widths = [35,14,20,22,12,35,12,18,30,30,8,20,18]

        for col_i, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col_i, value=h)
            cell.font = Font(name="Calibri", bold=True, color="E6EDF3", size=10)
            cell.fill = PatternFill("solid", fgColor="1C2128")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[ws.cell(row=1, column=col_i).column_letter].width = w

        for row_i, row in enumerate(rows, 2):
            for col_i, val in enumerate(row, 1):
                cell = ws.cell(row=row_i, column=col_i)
                if isinstance(val, datetime):
                    cell.value = val.strftime("%Y-%m-%d")
                else:
                    cell.value = val
                cell.fill = PatternFill("solid", fgColor="161B22" if row_i % 2 == 0 else "1C2128")
                cell.font = Font(name="Calibri", color="E6EDF3", size=9)
                if col_i == 11 and val is not None:
                    cell.font = Font(name="Calibri", bold=True, color=_score_color(val), size=9)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"prospects_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={fname}"})

    except Exception as e:
        logger.error(f"Export Excel error: {e}")
        return StreamingResponse(io.BytesIO(b"Error"), media_type="text/plain")


@router.get("/prospects.csv")
async def export_prospects_csv(
    region: Optional[str] = Query(None),
    limit: int = Query(10000, le=50000),
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    where, params = [], {"limit": limit}
    if region:
        where.append("p.region = :region"); params["region"] = region
    where_sql = "WHERE " + " AND ".join(where) if where else ""

    rows = (await db.execute(text(f"""
        SELECT p.company_name, p.siren, p.city, p.region, p.naf_code, p.naf_label,
               p.employee_range, p.phone, p.email, p.website, p.propensity_score,
               COALESCE(ps.name, 'Nouveau'), p.created_at
        FROM prospects p
        LEFT JOIN pipeline_stages ps ON p.stage_id = ps.id
        {where_sql}
        ORDER BY p.propensity_score DESC NULLS LAST LIMIT :limit
    """), params)).fetchall()

    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["Entreprise","SIREN","Ville","Région","Code NAF","Secteur","Effectif","Téléphone","Email","Site web","Score","Étape","Créé le"])
    for row in rows:
        w.writerow([str(v) if v is not None else "" for v in row])
    output.seek(0)

    fname = f"prospects_{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/pipeline.xlsx")
async def export_pipeline_excel(
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        stages = (await db.execute(text("SELECT id, name, color FROM pipeline_stages ORDER BY \"order\""))).fetchall()
        wb = Workbook()
        wb.remove(wb.active)

        for stage_id, stage_name, stage_color in stages:
            ws = wb.create_sheet(title=stage_name[:31])
            rows = (await db.execute(text("""
                SELECT p.company_name, p.city, p.phone, p.email, p.website, p.propensity_score, p.created_at
                FROM prospects p WHERE p.stage_id = :sid ORDER BY p.propensity_score DESC NULLS LAST
            """), {"sid": stage_id})).fetchall()

            for ci, h in enumerate(["Entreprise","Ville","Téléphone","Email","Site web","Score","Créé le"], 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True, color="E6EDF3"); cell.fill = PatternFill("solid", fgColor="1C2128")
            for ri, row in enumerate(rows, 2):
                for ci, val in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")

        if not wb.sheetnames:
            wb.create_sheet("Vide")

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=pipeline.xlsx"})
    except Exception as e:
        logger.error(f"Pipeline export error: {e}")
        return StreamingResponse(io.BytesIO(b"Error"), media_type="text/plain")


@router.get("/full-report.xlsx")
async def export_full_report(
    current_user: CurrentUser = None,
    db: AsyncSession = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook(); ws = wb.active; ws.title = "Rapport"
        total = (await db.execute(text("SELECT COUNT(*) FROM prospects"))).scalar() or 0
        with_email = (await db.execute(text("SELECT COUNT(*) FROM prospects WHERE email IS NOT NULL"))).scalar() or 0
        with_phone = (await db.execute(text("SELECT COUNT(*) FROM prospects WHERE phone IS NOT NULL"))).scalar() or 0
        avg_score = (await db.execute(text("SELECT AVG(propensity_score) FROM prospects WHERE propensity_score IS NOT NULL"))).scalar()

        ws.cell(row=1, column=1, value="Rapport B2B Prospector").font = Font(bold=True, size=14)
        ws.cell(row=3, column=1, value="Total prospects"); ws.cell(row=3, column=2, value=total)
        ws.cell(row=4, column=1, value="Avec email"); ws.cell(row=4, column=2, value=with_email)
        ws.cell(row=5, column=1, value="Avec téléphone"); ws.cell(row=5, column=2, value=with_phone)
        ws.cell(row=6, column=1, value="Score moyen"); ws.cell(row=6, column=2, value=round(float(avg_score), 1) if avg_score else 0)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=rapport_complet.xlsx"})
    except Exception as e:
        return StreamingResponse(io.BytesIO(b"Error"), media_type="text/plain")
