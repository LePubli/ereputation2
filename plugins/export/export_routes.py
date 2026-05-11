"""
Plugin Export — Génération de rapports Excel professionnels
Endpoint : GET /export/prospects.xlsx
          GET /export/pipeline.xlsx
          GET /export/full-report.xlsx
"""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional, List
from datetime import datetime
import io, logging

from core.database import get_db
from core.auth import get_current_active_user
from models.database.user import User

logger = logging.getLogger(__name__)
router = APIRouter()


# ─── Styles constants ───
HEADER_FILL    = "1C2128"   # bg-card dark
HEADER_FONT    = "E6EDF3"   # text-primary
ACCENT_BLUE    = "2F81F7"
ACCENT_GREEN   = "3FB950"
ACCENT_ORANGE  = "D29922"
ACCENT_RED     = "F85149"
ACCENT_PURPLE  = "8B5CF6"
ROW_ALT        = "161B22"   # bg-secondary
ROW_NORMAL     = "1C2128"
BORDER_COLOR   = "30363D"


def _color_for_score(score: Optional[int]) -> str:
    if score is None: return "8B949E"
    if score >= 75: return ACCENT_GREEN
    if score >= 50: return ACCENT_BLUE
    if score >= 25: return ACCENT_ORANGE
    return ACCENT_RED


@router.get("/prospects.xlsx")
async def export_prospects_excel(
    ids: Optional[List[str]] = Query(None),
    region: Optional[str] = Query(None),
    score_min: Optional[int] = Query(None),
    limit: int = Query(5000, le=10000),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export prospects en Excel formaté"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils import get_column_letter

        # ── Build query ──
        where_clauses = []
        params: dict = {"limit": limit}

        if ids:
            placeholders = ", ".join(f":id{i}" for i, _ in enumerate(ids))
            where_clauses.append(f"id IN ({placeholders})")
            for i, id_ in enumerate(ids):
                params[f"id{i}"] = id_
        if region:
            where_clauses.append("region = :region")
            params["region"] = region
        if score_min is not None:
            where_clauses.append("score >= :score_min")
            params["score_min"] = score_min

        where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""
        rows = db.execute(text(f"""
            SELECT company_name, siren, city, region, naf_code, naf_label,
                   employee_count, phone, email, website, score,
                   pipeline_stage, created_at, linkedin_url
            FROM prospects {where_sql}
            ORDER BY score DESC NULLS LAST, company_name
            LIMIT :limit
        """), params).fetchall()

        # ── Build workbook ──
        wb = Workbook()
        ws = wb.active
        ws.title = "Prospects"
        ws.sheet_view.showGridLines = False

        # Background
        ws.sheet_properties.tabColor = "2F81F7"

        # Column definitions
        cols = [
            ("Entreprise", 35), ("SIREN", 14), ("Ville", 20),
            ("Région", 22), ("Code NAF", 12), ("Secteur d'activité", 35),
            ("Effectif", 10), ("Téléphone", 18), ("Email", 30),
            ("Site web", 30), ("Score", 8), ("Étape Pipeline", 20),
            ("Date ajout", 15), ("LinkedIn", 35),
        ]

        # ── Header row ──
        header_font = Font(name="Calibri", bold=True, color=HEADER_FONT, size=10)
        header_fill = PatternFill("solid", fgColor=HEADER_FILL)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color=BORDER_COLOR)
        header_border = Border(bottom=Side(style="medium", color=ACCENT_BLUE))

        ws.row_dimensions[1].height = 32

        for col_idx, (col_name, col_width) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # ── Data rows ──
        for row_idx, row in enumerate(rows, start=2):
            is_alt = (row_idx % 2 == 0)
            row_fill = PatternFill("solid", fgColor=ROW_ALT if is_alt else ROW_NORMAL)
            score_val = row[10]

            data = [
                row[0],   # company_name
                row[1],   # siren
                row[2],   # city
                row[3],   # region
                row[4],   # naf_code
                row[5],   # naf_label
                row[6],   # employee_count
                row[7],   # phone
                row[8],   # email
                row[9],   # website
                score_val,
                row[11],  # pipeline_stage
                row[12].strftime("%d/%m/%Y") if row[12] else "",  # created_at
                row[13],  # linkedin_url
            ]

            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.font = Font(name="Calibri", color=HEADER_FONT, size=9)
                cell.alignment = Alignment(vertical="center")
                cell.border = Border(bottom=Side(style="thin", color=BORDER_COLOR))

                # Special formatting
                if col_idx == 11 and score_val is not None:  # Score
                    score_color = _color_for_score(score_val)
                    cell.font = Font(name="Calibri", bold=True, color=score_color, size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                if col_idx in (9, 10, 14) and value:  # Email, website, linkedin
                    cell.font = Font(name="Calibri", color=ACCENT_BLUE, size=9, underline="single")
                    if col_idx == 9 and value:
                        cell.hyperlink = f"mailto:{value}"
                    elif value and str(value).startswith("http"):
                        cell.hyperlink = str(value)

            ws.row_dimensions[row_idx].height = 20

        # ── Summary section below data ──
        summary_row = len(rows) + 3
        ws.cell(row=summary_row, column=1, value="RÉCAPITULATIF").font = Font(name="Calibri", bold=True, color=ACCENT_BLUE, size=11)

        summaries = [
            ("Total prospects", len(rows)),
            ("Avec email", sum(1 for r in rows if r[8])),
            ("Avec téléphone", sum(1 for r in rows if r[7])),
            ("Score moyen", round(sum(r[10] for r in rows if r[10]) / max(1, len([r for r in rows if r[10]])), 1)),
            ("Score max", max((r[10] for r in rows if r[10]), default=0)),
        ]
        for i, (label, val) in enumerate(summaries):
            ws.cell(row=summary_row + 1 + i, column=1, value=label).font = Font(name="Calibri", color="8D96A0", size=9)
            cell = ws.cell(row=summary_row + 1 + i, column=2, value=val)
            cell.font = Font(name="Calibri", bold=True, color=HEADER_FONT, size=9)

        # ── Freeze pane ──
        ws.freeze_panes = "A2"

        # ── Auto-filter ──
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

        # ── Metadata ──
        wb.properties.title = "Prospects Export — B2B Prospector"
        wb.properties.creator = "Le Publicitaire"
        wb.properties.description = f"Export du {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"prospects_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Export Excel error: {e}")
        raise HTTPException(500, f"Erreur génération Excel: {str(e)}")


@router.get("/pipeline.xlsx")
async def export_pipeline_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export pipeline Kanban en Excel avec une feuille par étape"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        STAGES = [
            ("Nouveau", "8B949E"),
            ("Contacté", "2F81F7"),
            ("Qualifié", "8B5CF6"),
            ("Proposition", "D29922"),
            ("Négociation", "F97316"),
            ("Gagné", "3FB950"),
            ("Perdu", "F85149"),
        ]

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for stage_name, stage_color in STAGES:
            rows = db.execute(text("""
                SELECT company_name, city, region, phone, email, website, score, created_at
                FROM prospects
                WHERE COALESCE(pipeline_stage, 'Nouveau') = :stage
                ORDER BY score DESC NULLS LAST
            """), {"stage": stage_name}).fetchall()

            ws = wb.create_sheet(title=stage_name)
            ws.sheet_properties.tabColor = stage_color
            ws.sheet_view.showGridLines = False

            # Title
            ws.merge_cells("A1:H1")
            title_cell = ws["A1"]
            title_cell.value = f"{'='*3} {stage_name.upper()} — {len(rows)} prospects {'='*3}"
            title_cell.font = Font(name="Calibri", bold=True, color=stage_color, size=13)
            title_cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            # Headers
            headers = ["Entreprise", "Ville", "Région", "Téléphone", "Email", "Site web", "Score", "Date ajout"]
            widths =  [35,           18,       20,       18,          30,      30,          8,       15       ]
            for i, (h, w) in enumerate(zip(headers, widths), 1):
                cell = ws.cell(row=2, column=i, value=h)
                cell.font = Font(name="Calibri", bold=True, color="E6EDF3", size=9)
                cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=Side(style="medium", color=stage_color))
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.row_dimensions[2].height = 22

            for r_idx, row in enumerate(rows, 3):
                alt = r_idx % 2 == 0
                row_fill = PatternFill("solid", fgColor=ROW_ALT if alt else ROW_NORMAL)
                data = [row[0], row[1], row[2], row[3], row[4], row[5], row[6],
                        row[7].strftime("%d/%m/%Y") if row[7] else ""]
                for c_idx, val in enumerate(data, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.fill = row_fill
                    cell.font = Font(name="Calibri", color="E6EDF3", size=9)
                    cell.border = Border(bottom=Side(style="thin", color=BORDER_COLOR))
                    cell.alignment = Alignment(vertical="center")
                    if c_idx == 7 and val is not None:
                        cell.font = Font(name="Calibri", bold=True, color=_color_for_score(val), size=10)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[r_idx].height = 18

            ws.freeze_panes = "A3"

        # Summary sheet
        ws_sum = wb.create_sheet(title="📊 Résumé", index=0)
        ws_sum.sheet_view.showGridLines = False
        ws_sum.sheet_properties.tabColor = "2F81F7"

        ws_sum.merge_cells("A1:C1")
        ws_sum["A1"].value = "RÉSUMÉ PIPELINE — B2B Prospector"
        ws_sum["A1"].font = Font(name="Calibri", bold=True, color="2F81F7", size=14)
        ws_sum["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
        ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 36
        ws_sum.column_dimensions["A"].width = 25
        ws_sum.column_dimensions["B"].width = 14
        ws_sum.column_dimensions["C"].width = 14

        for i, h in enumerate(["Étape", "Nombre", "% du total"], 1):
            c = ws_sum.cell(row=2, column=i, value=h)
            c.font = Font(name="Calibri", bold=True, color="E6EDF3", size=9)
            c.fill = PatternFill("solid", fgColor=HEADER_FILL)
            c.alignment = Alignment(horizontal="center")
        ws_sum.row_dimensions[2].height = 22

        total_all = db.execute(text("SELECT COUNT(*) FROM prospects")).scalar() or 1
        for r_i, (stage_name, stage_color) in enumerate(STAGES, 3):
            cnt = db.execute(text("SELECT COUNT(*) FROM prospects WHERE COALESCE(pipeline_stage,'Nouveau') = :s"), {"s": stage_name}).scalar() or 0
            row_fill = PatternFill("solid", fgColor=ROW_ALT if r_i % 2 == 0 else ROW_NORMAL)

            c1 = ws_sum.cell(row=r_i, column=1, value=stage_name)
            c1.font = Font(name="Calibri", bold=True, color=stage_color, size=10)
            c1.fill = row_fill

            c2 = ws_sum.cell(row=r_i, column=2, value=cnt)
            c2.font = Font(name="Calibri", color="E6EDF3", size=10)
            c2.alignment = Alignment(horizontal="center")
            c2.fill = row_fill

            c3 = ws_sum.cell(row=r_i, column=3, value=f"=B{r_i}/SUM(B3:B9)*100")
            c3.number_format = "0.0\"%\""
            c3.font = Font(name="Calibri", color="8D96A0", size=9)
            c3.alignment = Alignment(horizontal="center")
            c3.fill = row_fill

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"pipeline_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"})

    except Exception as e:
        logger.error(f"Pipeline export error: {e}")
        raise HTTPException(500, str(e))


@router.get("/full-report.xlsx")
async def export_full_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Rapport complet : prospects + pipeline + analytics + top régions"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.series import DataPoint

        wb = Workbook()
        wb.remove(wb.active)

        # ── Sheet 1: KPIs ──
        ws_kpi = wb.create_sheet("📊 KPIs")
        ws_kpi.sheet_view.showGridLines = False
        ws_kpi.sheet_properties.tabColor = "2F81F7"
        ws_kpi.column_dimensions["A"].width = 30
        ws_kpi.column_dimensions["B"].width = 18

        ws_kpi.merge_cells("A1:B1")
        ws_kpi["A1"].value = f"RAPPORT B2B PROSPECTOR — {datetime.now().strftime('%d/%m/%Y')}"
        ws_kpi["A1"].font = Font(name="Calibri", bold=True, color="2F81F7", size=14)
        ws_kpi["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
        ws_kpi["A1"].alignment = Alignment(horizontal="center")
        ws_kpi.row_dimensions[1].height = 36

        kpi_data_rows = [
            ("Indicateur", "Valeur", True),
            ("Total prospects", db.execute(text("SELECT COUNT(*) FROM prospects")).scalar() or 0, False),
            ("Avec email", db.execute(text("SELECT COUNT(*) FROM prospects WHERE email IS NOT NULL")).scalar() or 0, False),
            ("Avec téléphone", db.execute(text("SELECT COUNT(*) FROM prospects WHERE phone IS NOT NULL")).scalar() or 0, False),
            ("Score moyen", db.execute(text("SELECT ROUND(AVG(score)::numeric, 1) FROM prospects WHERE score IS NOT NULL")).scalar() or 0, False),
            ("Leads gagnés", db.execute(text("SELECT COUNT(*) FROM prospects WHERE pipeline_stage='Gagné'")).scalar() or 0, False),
            ("En négociation", db.execute(text("SELECT COUNT(*) FROM prospects WHERE pipeline_stage='Négociation'")).scalar() or 0, False),
        ]

        for r_i, (label, val, is_header) in enumerate(kpi_data_rows, 2):
            fill = PatternFill("solid", fgColor=HEADER_FILL if is_header else (ROW_ALT if r_i % 2 == 0 else ROW_NORMAL))
            for c_i, v in enumerate([label, val], 1):
                cell = ws_kpi.cell(row=r_i, column=c_i, value=v)
                cell.fill = fill
                cell.font = Font(name="Calibri", bold=is_header, color="E6EDF3" if is_header else ("2F81F7" if c_i == 2 else "E6EDF3"), size=10 if is_header else 9)
                cell.alignment = Alignment(vertical="center", horizontal="center" if c_i == 2 else "left")
                cell.border = Border(bottom=Side(style="thin", color=BORDER_COLOR))
            ws_kpi.row_dimensions[r_i].height = 22

        # ── Sheet 2: Top Régions with chart ──
        ws_reg = wb.create_sheet("📍 Régions")
        ws_reg.sheet_view.showGridLines = False
        ws_reg.sheet_properties.tabColor = "3FB950"
        ws_reg.column_dimensions["A"].width = 30
        ws_reg.column_dimensions["B"].width = 15

        regions = db.execute(text("""
            SELECT region, COUNT(*) as cnt FROM prospects
            WHERE region IS NOT NULL GROUP BY region ORDER BY cnt DESC LIMIT 12
        """)).fetchall()

        ws_reg["A1"].value = "RÉPARTITION PAR RÉGION"
        ws_reg["A1"].font = Font(name="Calibri", bold=True, color="3FB950", size=12)
        ws_reg["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
        ws_reg.row_dimensions[1].height = 28

        ws_reg.cell(row=2, column=1, value="Région").font = Font(name="Calibri", bold=True, color="E6EDF3")
        ws_reg.cell(row=2, column=2, value="Nombre").font = Font(name="Calibri", bold=True, color="E6EDF3")
        for c in ws_reg[2]: c.fill = PatternFill("solid", fgColor=HEADER_FILL)

        for i, (reg, cnt) in enumerate(regions, 3):
            fill = PatternFill("solid", fgColor=ROW_ALT if i % 2 == 0 else ROW_NORMAL)
            ws_reg.cell(row=i, column=1, value=reg).fill = fill
            ws_reg.cell(row=i, column=1).font = Font(name="Calibri", color="E6EDF3", size=9)
            ws_reg.cell(row=i, column=2, value=cnt).fill = fill
            ws_reg.cell(row=i, column=2).font = Font(name="Calibri", color="3FB950", bold=True, size=9)
            ws_reg.cell(row=i, column=2).alignment = Alignment(horizontal="center")

        # Bar chart for regions
        if regions:
            chart = BarChart()
            chart.type = "bar"
            chart.title = "Prospects par région"
            chart.style = 10
            chart.grouping = "clustered"
            data = Reference(ws_reg, min_col=2, min_row=2, max_row=2 + len(regions))
            cats = Reference(ws_reg, min_col=1, min_row=3, max_row=2 + len(regions))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            chart.width = 20
            chart.height = 14
            ws_reg.add_chart(chart, "D2")

        # ── Sheet 3: Prospects (top 1000) ──
        ws_pros = wb.create_sheet("🏢 Prospects")
        ws_pros.sheet_view.showGridLines = False
        ws_pros.sheet_properties.tabColor = "8B5CF6"

        all_rows = db.execute(text("""
            SELECT company_name, city, region, naf_label, employee_count,
                   phone, email, website, score, pipeline_stage
            FROM prospects ORDER BY score DESC NULLS LAST LIMIT 1000
        """)).fetchall()

        headers_p = ["Entreprise","Ville","Région","Secteur","Effectif","Tél","Email","Site","Score","Pipeline"]
        widths_p  = [32,          16,     18,      28,        9,        16,   26,    26,    8,     18   ]
        for i, (h, w) in enumerate(zip(headers_p, widths_p), 1):
            c = ws_pros.cell(row=1, column=i, value=h)
            c.font = Font(name="Calibri", bold=True, color="E6EDF3", size=9)
            c.fill = PatternFill("solid", fgColor=HEADER_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(bottom=Side(style="medium", color=ACCENT_BLUE))
            ws_pros.column_dimensions[get_column_letter(i)].width = w
        ws_pros.row_dimensions[1].height = 22

        for r_i, row in enumerate(all_rows, 2):
            alt = r_i % 2 == 0
            rfill = PatternFill("solid", fgColor=ROW_ALT if alt else ROW_NORMAL)
            for c_i, val in enumerate(row, 1):
                cell = ws_pros.cell(row=r_i, column=c_i, value=val)
                cell.fill = rfill
                cell.font = Font(name="Calibri", color="E6EDF3", size=8)
                cell.border = Border(bottom=Side(style="thin", color=BORDER_COLOR))
                cell.alignment = Alignment(vertical="center")
                if c_i == 9:  # Score
                    cell.font = Font(name="Calibri", bold=True, color=_color_for_score(val), size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_pros.row_dimensions[r_i].height = 16

        ws_pros.freeze_panes = "A2"
        ws_pros.auto_filter.ref = f"A1:{get_column_letter(len(headers_p))}1"

        # ── Save ──
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"rapport_complet_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"})

    except Exception as e:
        logger.error(f"Full report error: {e}")
        raise HTTPException(500, str(e))
